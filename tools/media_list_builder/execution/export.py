"""
Excel Exporter for Media List Builder
=======================================
Produces a formatted Excel spreadsheet with the media pitch list.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


COLUMNS = [
    ("First Name", 14),
    ("Last Name", 16),
    ("Outlet", 24),
    ("Website", 24),
    ("Role", 24),
    ("Media Type", 14),
    ("Location", 18),
    ("Pitch Angle", 40),
    ("Previous Story", 40),
    ("Story URL", 40),
    ("Email", 30),
    ("Notes", 30),
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def export_xlsx(result: dict, output_path: str):
    """Export media list to a formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Media List"

    # Freeze header row
    ws.freeze_panes = "A2"

    # Header row
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data rows
    for row_idx, contact in enumerate(result.get("contacts", []), 2):
        prev_title = contact.get("previous_story_title", "")
        prev_url = contact.get("previous_story_url", "")

        row_data = [
            contact.get("first_name", ""),
            contact.get("last_name", ""),
            contact.get("outlet", ""),
            contact.get("outlet_website", ""),
            contact.get("role", ""),
            contact.get("media_type", ""),
            contact.get("location", ""),
            contact.get("pitch_angle", ""),
            prev_title,
            prev_url,
            contact.get("email", ""),
            contact.get("notes", ""),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(result.get('contacts', [])) + 1}"

    # Set row height for data rows
    for row_idx in range(2, len(result.get("contacts", [])) + 2):
        ws.row_dimensions[row_idx].height = 30

    # Add metadata sheet
    meta_ws = wb.create_sheet("Info")
    meta_data = [
        ("Field", "Value"),
        ("Issue", result.get("issue", "")),
        ("Location", result.get("location", "")),
        ("Total Contacts", str(len(result.get("contacts", [])))),
        ("Pitch Timing", result.get("pitch_timing", "")),
    ]
    for row_idx, (key, val) in enumerate(meta_data, 1):
        meta_ws.cell(row=row_idx, column=1, value=key).font = Font(name="Calibri", bold=True)
        meta_ws.cell(row=row_idx, column=2, value=val).font = Font(name="Calibri")
    meta_ws.column_dimensions["A"].width = 18
    meta_ws.column_dimensions["B"].width = 60

    wb.save(output_path)
